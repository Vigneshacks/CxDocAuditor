"""
writer.py

Creates Excel reports from the scanned inventory.
"""

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook

from utils.logger import logger
from excel.formatter import format_workbook
from excel.summary import build_summary

def generate_excel_report(
    inventory: list[dict],
    output_folder: Path,
) -> Path:
    """
    Generate an Excel report from the scanned inventory.

    Parameters
    ----------
    inventory : list[dict]
        Inventory returned by scan_system()

    output_folder : Path
        Folder where the report will be saved.

    Returns
    -------
    Path
        Path to the generated workbook.
    """

    output_folder.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()

    inventory_sheet = workbook.active
    inventory_sheet.title = "Inventory"

    summary_sheet = workbook.create_sheet("Summary")

    headers = _build_headers(inventory)

    _write_headers(inventory_sheet, headers)

    _write_inventory(
        inventory_sheet,
        inventory,
        headers,
                    )

    summary = build_summary(inventory)

    _write_summary_sheet(
        summary_sheet,
        summary,
                    )
    output_path = _create_output_path(output_folder)
    format_workbook(workbook)

    workbook.save(output_path)

    # logger.info("Excel report saved to %s", output_path)

    return output_path


# ------------------------------------------------------------------------
# Helper Functions
# ------------------------------------------------------------------------


def _create_output_path(output_folder: Path) -> Path:
    """
    Create a timestamped Excel filename.
    """

    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")

    return output_folder / f"CxDoc_Report_{timestamp}.xlsx"


def _build_headers(inventory: list[dict]) -> list[str]:
    """
    Build Excel headers dynamically based on discovered levels.
    """

    headers = ["Equipment Reference"]

    levels = set()

    for equipment in inventory:

        levels.update(equipment["levels"].keys())

    for level in sorted(levels):

        if level == "L2-A":

            headers.extend(["FAT", "MIR", "WIR"])

        else:

            suffix = level.replace("-", "")

            headers.extend([
                f"FAT-{suffix}",
                f"MIR-{suffix}",
                f"WIR-{suffix}",
            ])

    return headers


def _write_headers(worksheet, headers):

    for column, header in enumerate(headers, start=1):

        worksheet.cell(
            row=1,
            column=column,
            value=header,
        )


def _write_inventory(
    worksheet,
    inventory,
    headers,
):
    """
    Write inventory rows into Excel.
    """

    # Build a lookup of column name -> Excel column number
    column_lookup = {
        header: index + 1
        for index, header in enumerate(headers)
    }

    row = 2

    for equipment in inventory:

        cell=worksheet.cell(
            row=row,
            column=1,
            value=equipment["reference"],
        )
        cell.hyperlink = str(equipment["path"])
        cell.style = "Hyperlink"

        for level_name, document_types in equipment["levels"].items():

            if level_name == "L2-A":

                mapping = {
                    "FAT": "FAT",
                    "MIR": "MIR",
                    "WIR": "WIR",
                }

            else:

                suffix = level_name.replace("-", "")

                mapping = {
                    "FAT": f"FAT-{suffix}",
                    "MIR": f"MIR-{suffix}",
                    "WIR": f"WIR-{suffix}",
                }

            for doc_type, info in document_types.items():

                documents = info["documents"]

                folder = info["folder"]
                header = mapping.get(doc_type)

                if header is None:
                    continue

                column = column_lookup[header]

                cell=worksheet.cell(
                    row=row,
                    column=column,
                    value="\n".join(f"• {doc}" for doc in documents),
                )
                cell.hyperlink = str(folder)
                cell.style = "Hyperlink"

        row += 1


def _write_summary_sheet(
    worksheet,
    summary,
):
    """
    Write summary sheet.
    """

    row = 1

    worksheet.cell(row=row, column=1, value="CxDoc Auditor Report")

    row += 2

    worksheet.cell(row=row, column=1, value="Equipment Scanned")
    worksheet.cell(row=row, column=2,
                   value=summary["equipment_count"])

    row += 3

    worksheet.cell(row=row, column=1, value="Level")
    worksheet.cell(row=row, column=2, value="FAT")
    worksheet.cell(row=row, column=3, value="MIR")
    worksheet.cell(row=row, column=4, value="WIR")

    row += 1

    for level in sorted(summary["levels"]):

        stats = summary["levels"][level]

        worksheet.cell(row=row, column=1, value=level)

        worksheet.cell(
            row=row,
            column=2,
            value=f'{stats["FAT"]} / {summary["equipment_count"]}'
        )

        worksheet.cell(
            row=row,
            column=3,
            value=f'{stats["MIR"]} / {summary["equipment_count"]}'
        )

        worksheet.cell(
            row=row,
            column=4,
            value=f'{stats["WIR"]} / {summary["equipment_count"]}'
        )

        row += 1

    row += 2

    worksheet.cell(row=row, column=1, value="Missing Documents")

    row += 1

    worksheet.cell(row=row, column=1, value="Level")
    worksheet.cell(row=row, column=2, value="Missing FAT")
    worksheet.cell(row=row, column=3, value="Missing MIR")
    worksheet.cell(row=row, column=4, value="Missing WIR")

    row += 1

    for level in sorted(summary["levels"]):

        stats = summary["levels"][level]

        worksheet.cell(row=row, column=1, value=level)
        worksheet.cell(row=row, column=2,
                       value=stats["Missing FAT"])
        worksheet.cell(row=row, column=3,
                       value=stats["Missing MIR"])
        worksheet.cell(row=row, column=4,
                       value=stats["Missing WIR"])

        row += 1