"""
formatter.py

Formatting utilities for CxDoc Auditor Excel reports.
"""

from openpyxl.styles import (
    Font,
    PatternFill,
    Border,
    Side,
    Alignment,
)
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------
# Styles
# ---------------------------------------------------------------------

HEADER_FILL = PatternFill(
    fill_type="solid",
    fgColor="1F4E78",
)

HEADER_FONT = Font(
    color="FFFFFF",
    bold=True,
    size=11,
)

MISSING_FILL = PatternFill(
    fill_type="solid",
    fgColor="F4CCCC",
)

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


# ---------------------------------------------------------------------
# Main Formatting Function
# ---------------------------------------------------------------------

def format_workbook(workbook):
    """
    Apply all formatting to the workbook.
    """

    worksheet = workbook.active

    format_headers(worksheet)
    auto_fit_columns(worksheet)
    wrap_text(worksheet)
    freeze_header(worksheet)
    add_filters(worksheet)
    adjust_row_height(worksheet)
    highlight_missing(worksheet)


# ---------------------------------------------------------------------
# Header Formatting
# ---------------------------------------------------------------------

def format_headers(worksheet):
    """
    Format the header row.
    """

    for cell in worksheet[1]:

        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )


# ---------------------------------------------------------------------
# Auto Column Width
# ---------------------------------------------------------------------

def auto_fit_columns(worksheet):
    """
    Automatically resize columns.
    """

    for column in worksheet.columns:

        max_length = 0

        column_letter = get_column_letter(column[0].column)

        for cell in column:

            if cell.value is not None:

                text = str(cell.value)

                longest_line = max(text.split("\n"), key=len)

                if len(longest_line) > max_length:
                    max_length = len(longest_line)

        worksheet.column_dimensions[column_letter].width = min(max_length + 3, 50)


# ---------------------------------------------------------------------
# Wrap Text
# ---------------------------------------------------------------------

def wrap_text(worksheet):
    """
    Enable text wrapping.
    """

    for row in worksheet.iter_rows():

        for cell in row:

            cell.alignment = Alignment(
                wrap_text=True,
                vertical="top",
            )


# ---------------------------------------------------------------------
# Freeze Header
# ---------------------------------------------------------------------

def freeze_header(worksheet):
    """
    Freeze first row.
    """

    worksheet.freeze_panes = "A2"


# ---------------------------------------------------------------------
# Auto Filter
# ---------------------------------------------------------------------

def add_filters(worksheet):
    """
    Enable filters.
    """

    worksheet.auto_filter.ref = worksheet.dimensions


# ---------------------------------------------------------------------
# Row Height
# ---------------------------------------------------------------------

def adjust_row_height(worksheet):
    """
    Increase row height for wrapped text.
    """

    for row in worksheet.iter_rows(min_row=2):

        worksheet.row_dimensions[row[0].row].height = 35


# ---------------------------------------------------------------------
# Highlight Missing Cells
# ---------------------------------------------------------------------

def highlight_missing(worksheet):
    """
    Highlight empty document cells.
    """

    for row in worksheet.iter_rows(min_row=2):

        # Skip Equipment Reference column
        for cell in row[1:]:

            if cell.value is None or str(cell.value).strip() == "":

                cell.fill = MISSING_FILL

            cell.border = THIN_BORDER


# ---------------------------------------------------------------------
# Center Headers Again
# ---------------------------------------------------------------------

def center_headers(worksheet):
    """
    Center align header row.
    """

    for cell in worksheet[1]:

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )